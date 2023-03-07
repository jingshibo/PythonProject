##  Automatic Recognition
import cv2
import easyocr
import os
import re
from openpyxl import load_workbook
from openpyxl.drawing.image import Image


## 在这里建立照片地址和学校名称对应关系的列表。注意严格按照如下格式，所有标点符号必须使用英文标点，而不是中文标点
school_name_list = {
    '深圳市宝安中学初中部': '深圳市宝安中学（集团）初中部',
    '深圳市宝安中学高中部': '深圳市宝安中学（集团）高中部',
    '深圳市宝安区人民政府': '深圳市宝安区人民政府',
}


## get the time automatically from the image
def getTimeAuto(image_folder_path):
    # Create an EasyOCR reader object with the desired languages
    reader = easyocr.Reader(['ch_sim', 'en'])
    # Keep track of the filenames that have already been processed
    processed_files = set()

    # Iterate over all the files in the folder
    for filename in os.listdir(image_folder_path):
        # default value
        text = '_未能正确识别地址'
        time = '_未能正确识别时间'
        # Check if the file is an image
        if filename.endswith('.jpg') or filename.endswith('.png') or filename.endswith('.jpeg') and filename not in processed_files:
            # Load the image
            image = cv2.imread(os.path.join(image_folder_path, filename))
            # Convert the image to grayscale
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            # Threshold the grayscale image to create a binary mask of the watermark
            _, mask = cv2.threshold(gray, 220, 255, cv2.THRESH_BINARY)
            # Apply the mask to the original image to extract the watermark
            watermark = cv2.bitwise_and(image, image, mask=mask)
            # Extract the text from the image using EasyOCR
            results = reader.readtext(watermark)

            # Process the text to get time and position information
            for result in results:
                if '深圳市' in result[1] and len(result[1] >= 7):  # address
                    text = result[1].replace("深圳市", "")
                    if (':' or '.') in text:
                        text = text.replace(":", "")
                elif re.match(r".*\d:\d{2}.*", result[1]):  # time
                    time = result[1].replace(":", "-")

            # Rename the image file using the recognized text
            _, extension = os.path.splitext(filename)
            new_filename = '{}_{}{}'.format(text, time, extension)
            new_filepath = os.path.join(image_folder_path, new_filename)
            # os.rename(os.path.join(folder_path, filename), os.path.join(folder_path, new_filename))

            # Check if the new filename already exists, and add a suffix to make it unique
            suffix = 1
            while os.path.exists(new_filepath):
                new_filename = '{}_{}_{}{}'.format(text, time, suffix, extension)
                new_filepath = os.path.join(image_folder_path, new_filename)
                suffix += 1

            # Rename the file
            os.rename(os.path.join(image_folder_path, filename), new_filepath)
            # Add the new filename to the set of processed files
            processed_files.add(new_filename)


## get the time manually, either '上班' or ’下班
def getTimeManually(image_folder_path, time):
    # Create an EasyOCR reader object with the desired languages
    reader = easyocr.Reader(['ch_sim', 'en'])
    # Keep track of the filenames that have already been processed
    processed_files = set()

    # Iterate over all the files in the folder
    for filename in os.listdir(image_folder_path):
        # Check if the file is an image
        if filename.endswith('.jpg') or filename.endswith('.png') or filename.endswith('.jpeg') and filename not in processed_files:
            # Load the image
            image = cv2.imread(os.path.join(image_folder_path, filename))
            # Convert the image to grayscale
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            # Threshold the grayscale image to create a binary mask of the watermark
            _, mask = cv2.threshold(gray, 220, 255, cv2.THRESH_BINARY)
            # Apply the mask to the original image to extract the watermark
            watermark = cv2.bitwise_and(image, image, mask=mask)
            # Extract the text from the image using EasyOCR
            results = reader.readtext(watermark)
            # Process only the string with '深圳市:' inside
            for result in results:
                if '深圳市' in result[1]:
                    text = result[1]
                    word_to_remove = "深圳市:"
                    text = text.replace(word_to_remove, "")

            # Rename the image file using the recognized text
            _, extension = os.path.splitext(filename)
            new_filename = '{}{}{}'.format(text, time, extension)
            new_filepath = os.path.join(image_folder_path, new_filename)

            # Check if the new filename already exists, and add a suffix to make it unique
            suffix = 1
            while os.path.exists(new_filepath):
                new_filename = '{}{}_{}{}'.format(text, time, suffix, extension)
                new_filepath = os.path.join(image_folder_path, new_filename)
                suffix += 1

            # Rename the file
            os.rename(os.path.join(image_folder_path, filename), new_filepath)
            # Add the new filename to the set of processed files
            processed_files.add(new_filename)


##  insert images into Excel in a alphabetical order of name
def insertImageAlphaOrder(excel_folder_path, image_folder_path, excel_filename, sheet_name, starting_row_number, image_column_number,
        img_width, img_height):
    # define the starting row index to insert
    row_index = starting_row_number
    # define the column index to insert
    column_index = image_column_number
    # Set the path of the excel file
    excel_path = os.path.join(excel_folder_path, excel_filename)
    # Load the workbook
    wb = load_workbook(excel_path)
    # Select the worksheet by name or index
    ws = wb[sheet_name]
    # Get a list of image filenames in the folder and sort them in a alphabetical order of name
    image_files = sorted([f for f in os.listdir(image_folder_path) if f.endswith('.jpg') or f.endswith('.png') or f.endswith('.jpeg')])

    # Iterate over all the sorted image files
    for i, image_filename in enumerate(image_files):
        # Check if the file is an image
        if image_filename.endswith('.jpg') or image_filename.endswith('.png') or image_filename.endswith('.jpeg'):
            # Load the image file
            img_path = os.path.join(image_folder_path, image_filename)
            # Create an image object from the file
            img = Image(img_path)
            # Set the dimensions of the image cell
            img.width = img_width
            img.height = img_height
            # Add the image to the worksheet
            ws.add_image(img, '{}{}'.format(column_index, row_index))
            # Get the required row height based on the image dimensions and set the row height accordingly
            ws.row_dimensions[row_index].height = img.height
            row_index += 1

    # Save the workbook
    wb.save(excel_path)


##  insert images into Excel when you have a table to convert the image name to the school name
def insertImageSameName(excel_folder_path, image_folder_path, excel_filename, sheet_name, image_column_index, school_column_index,
        img_width, img_height, school_name_list):
    # Set the path of the excel file
    excel_path = os.path.join(excel_folder_path, excel_filename)
    # Load the workbook
    wb = load_workbook(excel_path)
    # Select the worksheet by name or index
    ws = wb[sheet_name]
    # Get a list of image filenames in the folder and sort it in a alphabetical order of name
    image_files = sorted([f for f in os.listdir(image_folder_path) if f.endswith('.jpg') or f.endswith('.png') or f.endswith('.jpeg')])

    # Iterate over all the sorted image files
    for i, image_filename in enumerate(image_files):
        # Check if the file is an image
        if image_filename.endswith('.jpg') or image_filename.endswith('.png') or image_filename.endswith('.jpeg'):
            # Load the image file
            img_path = os.path.join(image_folder_path, image_filename)
            # Create an image object from the file
            img = Image(img_path)
            # Set the dimensions of the image cell
            img.width = img_width
            img.height = img_height

            # Find the cell with the same filename
            filename_without_extension = os.path.splitext(image_filename)[0]
            try:
                school_name = filename_without_extension.split("_")[0]
                filename_without_extension = school_name_list[school_name]  # find the corresponding school name from the dict
            except KeyError:
                print(f"没有查找到该学校名称: {school_name}")
                continue

            for row in ws.iter_rows():
                if row[school_column_index].value == filename_without_extension:  # the 0 here means the column number at this row
                    # Add the image to the worksheet
                    ws.add_image(img, '{}{}'.format(image_column_index, row[school_column_index].row))  # row[0].row returns the row number of the cell row[0]
                    # Calculate the required row height based on the image dimensions and set the row height accordingly
                    ws.row_dimensions[row[school_column_index].row].height = img.height

    # Save the workbook
    wb.save(excel_path)

