# %%  Automatic Recognition
import cv2
import easyocr
import os
import re
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import string
import random
import unicodedata

# %% 在这里建立照片地址和学校名称对应关系的列表。注意严格按照如下格式，所有标点符号必须使用英文标点，而不是中文标点
school_name_list = {'上合小学': '上合小学', '松岗实验学校': '松岗实验学校', '育贤学校': '育贤学校', '安乐小学': '安乐', '文山小学': '文山小学', }


# %% get the time automatically from the image
def getTimeAuto(image_folder_path):
    # Create an EasyOCR reader object with the desired languages
    reader = easyocr.Reader(['ch_sim', 'en'])
    # Keep track of the filenames that have already been processed
    processed_files = set()

    # Iterate over all the files in the folder
    recognized_number = 0
    total_number = 0
    for filename in os.listdir(image_folder_path):
        # default value
        text = '_未能识别地址'
        time = '_未能识别时间'

        # Check if the file is an image
        if filename.endswith('.jpg') or filename.endswith('.png') or filename.endswith('.jpeg') and filename not in processed_files:
            # if all characters in the filename are Chinese, not processing it because this means it has been proceed
            name, _ = os.path.splitext(filename)
            if all('CJK' in unicodedata.name(char) for char in name):
                continue

            # rename Chinese files
            filename = renameFileReplacingChinese(image_folder_path, filename)

            # the number of images in total
            total_number += 1
            # Load the image
            image = cv2.imread(os.path.join(image_folder_path, filename))
            # Convert the image to grayscale
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            # Threshold the grayscale image to create a binary mask of the watermark
            _, mask = cv2.threshold(gray, 215, 255, cv2.THRESH_BINARY)
            # Apply the mask to the original image to extract the watermark
            watermark = cv2.bitwise_and(image, image, mask=mask)
            # Extract the text from the image using EasyOCR
            results = reader.readtext(watermark)

            # Process the text to get time and position information
            for number, result in enumerate(results):
                # Process only the string with '深圳市:' inside
                if '深圳市' in result[1]:
                    if len(result[1]) > 3:  # the address is stored in the current cell
                        text = result[1]
                    elif len(result[1]) == 3:  # the address is stored in the next cell
                        text = results[number + 1][1]
                    # remove certain words
                    text = text.replace("深圳市", "")
                    if '小小' in text:
                        text = text.replace('小小', '小')
                    # remove non-Chinese characters
                    pattern = re.compile(r'[^\u4e00-\u9fff]')
                    text = re.sub(pattern, '', text)
                    break
                # Process the picturing time
                elif re.match(r".*\d:\d{2}.*", result[1]):
                    time = result[1].replace(":", "-")

            # count the number of photos recognized
            if text != '_未能识别地址' and time != '_未能识别时间':
                recognized_number += 1

            # Rename the image file using the recognized text
            _, extension = os.path.splitext(filename)
            new_filename = '{}_{}{}'.format(text, time, extension)
            new_filepath = os.path.join(image_folder_path, new_filename)
            # os.rename(os.path.join(folder_path, filename), os.path.join(folder_path, new_filename))

            # Check if the new filename already exists, and add a suffix to make it unique
            suffix = 1
            while os.path.exists(new_filepath):
                new_filename = '{}_{}_{}{}'.format(text, time, suffix, extension)
                new_filepath = os.path.join(image_folder_path, new_filename)
                suffix += 1
            if suffix > 1:
                print('存在名称重复的照片，已用数字后缀进行标识，请检查')

            # Rename the file
            os.rename(os.path.join(image_folder_path, filename), new_filepath)
            # Add the new filename to the set of processed files
            processed_files.add(new_filename)

    print(f'总共新增{total_number}张图片，识别出{recognized_number}张图片')
    unrecognized_number = total_number - recognized_number
    if unrecognized_number > 0:
        print(f'有{unrecognized_number}张图片未能识别出地址或时间，请检查图片并手动重命名图片为正确地址。')


# %% get the time manually, either '上班' or ’下班
def getTimeManually(image_folder_path, time):
    # Create an EasyOCR reader object with the desired languages
    reader = easyocr.Reader(['ch_sim', 'en'])
    # Keep track of the filenames that have already been processed
    processed_files = set()

    # Iterate over all the files in the folder
    recognized_number = 0
    total_number = 0
    # Iterate over all the files in the folder
    for filename in os.listdir(image_folder_path):
        # default value
        text = '_未能识别地址'

        # Check if the file is an image
        if filename.endswith('.jpg') or filename.endswith('.png') or filename.endswith('.jpeg') and filename not in processed_files:
            # if all characters in the filename are Chinese, not processing it because this means it has been proceed
            name, _ = os.path.splitext(filename)
            if all('CJK' in unicodedata.name(char) for char in name):
                continue

            # rename Chinese files
            filename = renameFileReplacingChinese(image_folder_path, filename)

            # the number of images in total
            total_number += 1
            # Load the image
            image = cv2.imread(os.path.join(image_folder_path, filename))
            # Convert the image to grayscale
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            # Threshold the grayscale image to create a binary mask of the watermark
            _, mask = cv2.threshold(gray, 215, 255, cv2.THRESH_BINARY)
            # Apply the mask to the original image to extract the watermark
            watermark = cv2.bitwise_and(image, image, mask=mask)
            # Extract the text from the image using EasyOCR
            results = reader.readtext(watermark)

            # Process only the string with '深圳市:' inside
            for number, result in enumerate(results):
                if '深圳市' in result[1]:
                    if len(result[1]) > 3:  # the address is stored in the current cell
                        text = result[1]
                    elif len(result[1]) == 3:  # the address is stored in the next cell
                        text = results[number + 1][1]
                    # remove certain words
                    text = text.replace("深圳市", "")
                    if '小小' in text:
                        text = text.replace('小小', '小')
                    # remove non-Chinese characters
                    pattern = re.compile(r'[^\u4e00-\u9fff]')
                    text = re.sub(pattern, '', text)
                    break

            # count the number of photos recognized
            if text != '_未能识别地址':
                recognized_number += 1

            # Rename the image file using the recognized text
            _, extension = os.path.splitext(filename)
            new_filename = '{}{}{}'.format(text, time, extension)
            new_filepath = os.path.join(image_folder_path, new_filename)

            # Check if the new filename already exists, and add a suffix to make it unique
            suffix = 1
            while os.path.exists(new_filepath):
                new_filename = '{}{}_{}{}'.format(text, time, suffix, extension)
                new_filepath = os.path.join(image_folder_path, new_filename)
                suffix += 1
            if suffix > 1:
                print('存在名称重复的照片，已用数字后缀进行标识，请检查')

            # Rename the file
            os.rename(os.path.join(image_folder_path, filename), new_filepath)
            # Add the new filename to the set of processed files
            processed_files.add(new_filename)

    print(f'总共新增{total_number}张图片，识别出{recognized_number}张图片')
    unrecognized_number = total_number - recognized_number
    if unrecognized_number > 0:
        print(f'有{unrecognized_number}张图片未能识别出地址，请检查图片并手动重命名图片为正确地址。')


# %%  insert images into Excel when you have a table to convert the image name to the school name
def insertImageSameName(excel_folder_path, image_folder_path, excel_filename, sheet_name, image_column_index, school_column_index,
        img_width, img_height, school_name_list):
    # Set the path of the excel file
    excel_path = os.path.join(excel_folder_path, excel_filename)
    # Load the workbook
    wb = load_workbook(excel_path)
    # Select the worksheet by name or index
    ws = wb[sheet_name]
    # Get a list of image filenames in the folder and sort it in a alphabetical order of name
    image_files = sorted([f for f in os.listdir(image_folder_path) if f.endswith('.jpg') or f.endswith('.png') or f.endswith('.jpeg')])

    # Iterate over all the sorted image files
    total_number = 0
    recognized_number = 0
    for i, image_filename in enumerate(image_files):
        # Check if the file is an image
        if image_filename.endswith('.jpg') or image_filename.endswith('.png') or image_filename.endswith('.jpeg'):
            # the number of images in total
            total_number += 1
            # Load the image file
            img_path = os.path.join(image_folder_path, image_filename)
            # Create an image object from the file
            img = Image(img_path)
            # Set the dimensions of the image cell
            img.width = img_width
            img.height = img_height

            # Find the cell with the same filename
            filename_without_extension = os.path.splitext(image_filename)[0]
            try:
                school_name = filename_without_extension.split("_")[0]
                filename_without_extension = school_name_list[school_name]  # find the corresponding school name from the dict
            except KeyError:
                print(f"没有在学校列表的左侧中找到图片上对应的学校: {school_name}")
                continue

            indicator = 0  # indicate if the new image is added
            for row in ws.iter_rows():
                if row[school_column_index].value == filename_without_extension:  # the 0 here means the column number at this row
                    # Add the image to the worksheet
                    ws.add_image(img, '{}{}'.format(image_column_index,
                        row[school_column_index].row))  # row[0].row returns the row number of the cell row[0]
                    # Calculate the required row height based on the image dimensions and set the row height accordingly
                    ws.row_dimensions[row[school_column_index].row].height = img.height
                    # count the number of photos recognized
                    recognized_number += 1
                    indicator = 1

            if indicator == 0:
                print(f"学校列表右侧中的'{filename_without_extension}'在excel表格上没有对应的学校")
    # Save the workbook
    wb.save(excel_path)

    print(f'总共有{total_number}张图片，插入了{recognized_number}张图片')
    unrecognized_number = total_number - recognized_number
    if unrecognized_number > 0:
        print(f'有{unrecognized_number}张图片未正确插入。(可能原因1：图片地址识别错误。可能原因2：这是一个新的地址。可能原因3：excel表上没有对应的学校)')


# %%  insert images into Excel in a alphabetical order of name
def insertImageAlphaOrder(excel_folder_path, image_folder_path, excel_filename, sheet_name, starting_row_number, image_column_number,
        img_width, img_height):
    # define the starting row index to insert
    row_index = starting_row_number
    # define the column index to insert
    column_index = image_column_number
    # Set the path of the excel file
    excel_path = os.path.join(excel_folder_path, excel_filename)
    # Load the workbook
    wb = load_workbook(excel_path)
    # Select the worksheet by name or index
    ws = wb[sheet_name]
    # Get a list of image filenames in the folder and sort them in a alphabetical order of name
    image_files = sorted([f for f in os.listdir(image_folder_path) if f.endswith('.jpg') or f.endswith('.png') or f.endswith('.jpeg')])

    # Iterate over all the sorted image files
    for i, image_filename in enumerate(image_files):
        # Check if the file is an image
        if image_filename.endswith('.jpg') or image_filename.endswith('.png') or image_filename.endswith('.jpeg'):
            # Load the image file
            img_path = os.path.join(image_folder_path, image_filename)
            # Create an image object from the file
            img = Image(img_path)
            # Set the dimensions of the image cell
            img.width = img_width
            img.height = img_height
            # Add the image to the worksheet
            ws.add_image(img, '{}{}'.format(column_index, row_index))
            # Get the required row height based on the image dimensions and set the row height accordingly
            ws.row_dimensions[row_index].height = img.height
            row_index += 1

    # Save the workbook
    wb.save(excel_path)


# %% find the files with Chinese Characters and replace the name with a random string
def renameFileReplacingChinese(image_folder_path, filename):
    # generate a random string for replacement
    def generate_random_string(length):
        letters = string.ascii_lowercase + string.ascii_uppercase + string.digits
        return ''.join(random.choice(letters) for i in range(length))

    random_string = generate_random_string(10)

    # Replace the Chinese characters in the filename with a random string
    def replace_chinese_with_a(file_name):
        pattern = re.compile(r'[\u4e00-\u9fff]+')
        has_chinese = bool(pattern.search(file_name))
        new_file_name = pattern.sub(random_string, file_name) if has_chinese else file_name
        return new_file_name

    # Rename the file with Chinese characters replaces
    new_filename = replace_chinese_with_a(filename)
    old_path = os.path.join(image_folder_path, filename)
    new_path = os.path.join(image_folder_path, new_filename)
    os.rename(old_path, new_path)

    return new_filename

