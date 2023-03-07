
## another way to realize result reorganization
# from collections import defaultdict
#
# # Recursive function to convert nested defaultdicts to dicts
# def defaultdict_to_dict(d):
#     if isinstance(d, defaultdict):
#         d = {k: defaultdict_to_dict(v) for k, v in d.items()}
#     return d
#
# # Create an empty dictionary to hold the combined extracted results
# combined_results = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(list))))
#
# # Iterate over all the keys in the extracted_results dictionary
# for subject_number, subject_results in extracted_results.items():
#     for dataset, datavalue in subject_results.items():
#         for condition, results in datavalue.items():
#             for item, value in results.items():
#                 # Iterate over each key-value pair in the results dictionary
#                 for extract_delay_key, extract_delay_value in value.items():
#                     # Add the extract_delay value for this item across all subject_number to the corresponding list
#                     combined_results[dataset][condition][item][extract_delay_key].append(extract_delay_value)
#
# # Convert each level of the defaultdict to a dict
# combined_results = defaultdict_to_dict(combined_results)


##
import cv2
import easyocr
import os
import re

# Load the image
image = cv2.imread('D:\Project\pythonProject\Test\-125009aec32cc5c5.jpg')

# Convert the image to grayscale
gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

# Threshold the grayscale image to create a binary mask of the watermark
_, mask = cv2.threshold(gray, 220, 255, cv2.THRESH_BINARY)

# Apply the mask to the original image to extract the watermark
watermark = cv2.bitwise_and(image, image, mask=mask)

# Save the extracted watermark as a new image file
cv2.imwrite('extracted_watermark.png', watermark)


## recognize the watermark
reader = easyocr.Reader(['ch_sim', 'en'])
results = reader.readtext(watermark)

for result in results:
    if '深圳市:' in result[1]:
        text = result[1]
        word_to_remove = "深圳市:"
        text = text.replace(word_to_remove, "")
    elif re.match(r".*\d:\d{2}.*", result[1]):
        time = result[1].replace(":", ".")


## Rename the image file using the recognized text
filename, extension = os.path.splitext('extracted_watermark.png')
new_filename = '{}_{}{}'.format(text, time, extension)
os.rename('extracted_watermark.png', new_filename)


##  insert images into Excel
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# define the column index of schools
school_column_index = 0

# define the column index to insert
image_column_index = 'C'

# define the image width
img_width = 80

# define the image height
img_height = 80

# Set the path to the folder containing the images
folder_path = 'D:\Project\pythonProject\Test'

# Set the name of the Excel file to read from
filename = 'Book12.xlsx'

# Set the path of the excel file
excel_path = os.path.join(folder_path, filename)

# Set the name of the sheet to insert the images into
sheet_name = 'Sheet1'

# Set the column letter to insert the images into
column_letter = image_column_index

# Load the workbook
wb = load_workbook(excel_path)

# Select the worksheet by name or index
ws = wb[sheet_name]

# Get a list of image filenames in the folder and sort it in a alphabetical order of name
image_files = sorted([f for f in os.listdir(folder_path) if f.endswith('.jpg') or f.endswith('.png') or f.endswith('.jpeg')])

# Iterate over all the sorted image files
for i, filename in enumerate(image_files):
    # Check if the file is an image
    if filename.endswith('.jpg') or filename.endswith('.png') or filename.endswith('.jpeg'):
        # Load the image file
        img_path = os.path.join(folder_path, filename)

        # Create an image object from the file
        img = Image(img_path)

        # Set the dimensions of the image cell
        img.width = 80
        img.height = 80

        # Find the cell with the matching filename
        filename_without_extension = os.path.splitext(filename)[0]
        for row in ws.iter_rows():
            if row[school_column_index].value == filename_without_extension:  # the 0 here means the column number at this row
                # Add the image to the worksheet
                ws.add_image(img, '{}{}'.format(image_column_index, row[school_column_index].row))  # row[0].row returns the row number of the cell row[0]
                # Calculate the required row height based on the image dimensions and set the row height accordingly
                ws.row_dimensions[row[school_column_index].row].height = img.height

# Save the workbook
wb.save(excel_path)


## Manually Recognition
import cv2
import easyocr
import os

# set the time
time = '上班'  # or '下班'

# Set the path to the folder containing the images
folder_path = 'D:\Project\pythonProject\Test'

# Create an EasyOCR reader object with the desired languages
reader = easyocr.Reader(['ch_sim', 'en'])

# Keep track of the filenames that have already been processed
processed_files = set()

# Iterate over all the files in the folder
for filename in os.listdir(folder_path):
    # Check if the file is an image
    if filename.endswith('.jpg') or filename.endswith('.png') or filename.endswith('.jpeg') and filename not in processed_files:
        # Load the image
        image = cv2.imread(os.path.join(folder_path, filename))

        # Convert the image to grayscale
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

        # Threshold the grayscale image to create a binary mask of the watermark
        _, mask = cv2.threshold(gray, 220, 255, cv2.THRESH_BINARY)

        # Apply the mask to the original image to extract the watermark
        watermark = cv2.bitwise_and(image, image, mask=mask)

        # Extract the text from the image using EasyOCR
        results = reader.readtext(watermark)

        # Process only the string with '深圳市:' inside
        for result in results:
            if '深圳市:' in result[1]:
                text = result[1]
                word_to_remove = "深圳市:"
                text = text.replace(word_to_remove, "")

        # Rename the image file using the recognized text
        _, extension = os.path.splitext(filename)
        new_filename = '{}_{}{}'.format(text, time, extension)
        new_filepath = os.path.join(folder_path, new_filename)
        # os.rename(os.path.join(folder_path, filename), os.path.join(folder_path, new_filename))

        # Check if the new filename already exists, and add a suffix to make it unique
        suffix = 1
        while os.path.exists(new_filepath):
            new_filename = '{}_{}_{}{}'.format(text, time, suffix, extension)
            new_filepath = os.path.join(folder_path, new_filename)
            suffix += 1

        # Rename the file
        os.rename(os.path.join(folder_path, filename), new_filepath)

        # Add the new filename to the set of processed files
        processed_files.add(new_filename)